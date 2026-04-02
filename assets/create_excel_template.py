#!/usr/bin/env python3
"""
create_excel_template.py
Creates assets/mokwon_template.xlsx with headers and sample data rows.
"""

import os
import sys

# Try openpyxl first, then fall back to xlwt/xlsxwriter or basic xlsx via zipfile
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'mokwon_template.xlsx')

HEADERS = [
    'student_id', 'full_name', 'gender', 'birth_year', 'birth_month', 'birth_day',
    'passport_number', 'place_of_birth', 'address', 'phone', 'school_name', 'agency_name',
    'sponsor_name', 'sponsor_address', 'sponsor_occupation', 'sponsor_relation', 'sponsor_contact_HP',
    'photo_path'
]

SAMPLE_ROWS = [
    [
        'STU001', 'Mohammed Rahman', 'M', '1999', '03', '15',
        'AA1234567', 'Dhaka', '123 Green Road, Dhaka', '+8801711234567',
        'Dhaka College', 'EduBridge Agency',
        'Abdul Rahman', '45 Mirpur, Dhaka', 'Business', 'Father', '+8801811234567',
        ''
    ],
    [
        'STU002', 'Fatima Begum', 'F', '2000', '07', '22',
        'BB9876543', 'Chittagong', '56 Port Road, Chittagong', '+8801912345678',
        'Chittagong High School', 'Global Studies',
        'Karim Begum', '78 Pahartali, Chittagong', 'Teacher', 'Mother', '+8801612345678',
        ''
    ]
]

if HAS_OPENPYXL:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write sample rows
    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Set column widths
    col_widths = [12, 22, 8, 12, 13, 11, 16, 18, 30, 18, 25, 22, 22, 30, 22, 18, 22, 30]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze first row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT_PATH)
    print(f'Created {OUTPUT_PATH} using openpyxl')

else:
    # Fallback: create minimal XLSX using zipfile + raw XML
    import zipfile
    import io

    def xml_escape(s):
        return (str(s)
                .replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;'))

    # Build shared strings
    all_strings = []
    string_index = {}

    def get_string_idx(s):
        s = str(s)
        if s not in string_index:
            string_index[s] = len(all_strings)
            all_strings.append(s)
        return string_index[s]

    # Build rows data
    all_rows = [HEADERS] + SAMPLE_ROWS

    # Pre-populate shared strings
    for row in all_rows:
        for cell in row:
            get_string_idx(cell)

    # sheet1.xml
    sheet_xml_parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        '<sheetData>',
    ]
    for row_idx, row_data in enumerate(all_rows, start=1):
        sheet_xml_parts.append(f'<row r="{row_idx}">')
        for col_idx, cell_val in enumerate(row_data):
            col_letter = chr(ord('A') + col_idx)
            s_idx = get_string_idx(cell_val)
            sheet_xml_parts.append(f'<c r="{col_letter}{row_idx}" t="s"><v>{s_idx}</v></c>')
        sheet_xml_parts.append('</row>')
    sheet_xml_parts.append('</sheetData></worksheet>')
    sheet_xml = ''.join(sheet_xml_parts)

    # sharedStrings.xml
    ss_xml_parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="{len(all_strings)}" uniqueCount="{len(all_strings)}">',
    ]
    for s in all_strings:
        ss_xml_parts.append(f'<si><t>{xml_escape(s)}</t></si>')
    ss_xml_parts.append('</sst>')
    ss_xml = ''.join(ss_xml_parts)

    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>'''

    rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''

    workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Students" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>'''

    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>'''

    with zipfile.ZipFile(OUTPUT_PATH, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types)
        zf.writestr('_rels/.rels', rels)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)
        zf.writestr('xl/sharedStrings.xml', ss_xml)

    print(f'Created {OUTPUT_PATH} using fallback ZIP method')

print('Done.')
